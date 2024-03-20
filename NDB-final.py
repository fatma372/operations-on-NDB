# -*- coding: utf-8 -*-
#true
import openpyxl

def PrefixAlgo(DB, length):
    W = set()
    for BitCount in range(1, length + 1):
        possible_Prefixs = [format(i, '0' + str(BitCount) + 'b') for i in range(2**BitCount)]
        for possible_Prefix in possible_Prefixs:
            if not any(s.startswith(possible_Prefix) for s in DB):
                new_element = possible_Prefix + '*' * (length - len(possible_Prefix))
                flag = True
                for ele in W:
                    tflag = False
                    for p, v in zip(ele, new_element):
                        if p != '*' and p != v:
                            tflag = True
                    if not tflag:
                        flag = False
                if flag:
                    W.add(new_element)
    return W

def StringToBinary(data):
    binary_data = []
    for item in data:
        binary_item = ''.join(format(ord(ch), '08b') for ch in item)
        binary_data.append(binary_item)
    return binary_data

def Search(NDB, element): 
    BinaryElement = ''.join(format(ord(ch), '08b') for ch in element)
    for item in NDB:
        for bit, char in zip(BinaryElement, item):
            if bit != char and char != '*':
                break
        else:
            return True
    return False

def Update(DB, old_element, new_element):
    if old_element in DB:
        index = DB.index(old_element)
        DB[index] = new_element

def Delete(DB, element):
    if element in DB:
        DB.remove(element)

def Insert(DB, element):
    DB.append(element)

# *************************************read data******************************************
file = "DB.xlsx"  
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

DB = []
for row in sheet.iter_rows(values_only=True):
    DB.extend(row)
DB_binary = StringToBinary(DB)

l = int(input("Enter the value of l: "))
print('******************************************************')
print("DB", DB)
print('_______________________________________________________')
print("\nDB in binary =", DB_binary)
print('_______________________________________________________')
print("\nl =", l)
print('_______________________________________________________')
NDB = PrefixAlgo(DB_binary, l)
print("\nNDB = ", NDB)
print('_______________________________________________________')


print("\n*********************   operations:  ********************")

print("\nSearch for an element in NDB:")
SearchElement = input("Enter the element to search for: ")
print("Founded in NDB :", Search(NDB, SearchElement))
print('>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>')

print("\nUpdate an element in DB:")
old_element = input("Enter the old element value: ")
new_element = input("Enter the new element value: ")
Update(DB, old_element, new_element)
for row in sheet.iter_rows():
    for cell in row:
        if cell.value == old_element:
            cell.value = new_element
workbook.save(file)
print("DB after update (Excel file):" ,DB)
print('>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>')


print("\nDelete an element from DB:")
DeleteElement = input("Enter the element to delete: ")
Delete(DB, DeleteElement)
for row in sheet.iter_rows():
    for cell in row:
        if cell.value == DeleteElement:
            cell.value = None
# Delete all cells with value None
for row in sheet:
    for cell in row:
        if cell.value is None:
            sheet.cell(row=cell.row, column=cell.column, value='*')
workbook.save(file)
print("DB after delete (Excel file):" ,DB)
print('>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>')


print("\nInsert an element into DB:")
InsertElement = input("Enter the element to insert: ")
Insert(DB, InsertElement)
sheet.append([InsertElement])
workbook.save(file)
print("DB after insert:" ,DB)

# **********************Update NDB*******************************
print('******************************************************************')

print("FINAL NDB ->\n")
DB = []
for row in sheet.iter_rows(values_only=True):
    DB.extend(row)
    
DB_binary_last = StringToBinary(DB)

final_NDB = PrefixAlgo(DB_binary_last, l)
print("NDB",final_NDB)
